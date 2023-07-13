#!/usr/bin/env python3

from openpyxl import load_workbook
import io, sys, requests

# Replace this with your OneDrive share link
onedrive_share_link = sys.argv[1]#'https://1drv.ms/x/s!Aexample_share_link'

# Convert OneDrive share link to direct download link
direct_link = onedrive_share_link.replace('https://1drv.ms', 'https://api.onedrive.com/v1.0/shares/u!')

# Add /root/content to the direct download link
direct_link += '/root/content'

# Download the Excel file
response = requests.get(direct_link)

# Check if the download was successful
if response.status_code == 200:
    # Load the workbook into memory
    workbook = load_workbook(io.BytesIO(response.content))
    
    # Access the first sheet (by default)
    sheet = workbook.active
    
    # Iterate through rows and cells
    for row in sheet.iter_rows():
        for cell in row:
            print(cell.value, end=' ')
        print()
else:
    print(f"Error: {response.status_code}{response.content}")
    